import sys
import json
import os
import openpyxl


def ReadDataFromJsonFile(type,value):
    #print("reading data from json file")
    ## Read Json file from Config directory
    cwd = os.getcwd()
    pcwd = "\\".join(cwd.split('\\')[:-1])
    #confdir = pcwd + "\\Config"
    confdir = cwd + "\\Config"
    json_data = open(confdir+"\\AutomationToolConfig.json")
    data = json.load(json_data)
   # print(data)
    try:
      return data[type][value]
    except:
      return None


def ReadDataFromRepo(type, name, value):
    get_input_type = ReadDataFromJsonFile("tool", "configfile")
    if get_input_type == "xls":
        strColValues = type+"|"+name
        try:
            return fnReadDataFromExcel("LogicalValues.xlsx", "LogicalNames", "Type|Name", strColValues, value)
        except:
            return None
    else:
        cwd = os.getcwd()
        pcwd = "\\".join(cwd.split('\\')[:-1])
        confdir = cwd + "\\Config"
        repo_file = open(confdir+"\\Repository.json")
        repo_data = json.load(repo_file)
        try:
            return repo_data[type][name][value]
        except:
            return None

def fnReadDataFromExcel(strExcelName,strSheetName,strUniqueColNames, strUniqueColValues, strValue):
        cwd = os.getcwd()
        pcwd = "\\".join(cwd.split('\\')[:-1])
        confdir = pcwd + "\\Config"
        wb = openpyxl.load_workbook(confdir+"\\" + strExcelName)
        sheet = wb.get_sheet_by_name(strSheetName)
        maxrows = sheet.max_row
        maxcols = sheet.max_column
        blnfound = False
        blnList = False

        ColValuesList = strUniqueColValues.split('|')
        if '|'  in strUniqueColValues:
         blnList = True

        for r in range(2,maxrows+1):
            counter = 0
            for c in range(1,maxcols+1):
                for i in range(0,len(ColValuesList)):
                    if sheet.cell(row=r,column=c).value == ColValuesList[i]:
                        counter+=1
                        if counter == len(ColValuesList) and blnList == True:
                            UniqueRowNo = r
                            #print("Unique Row number is : "+str(UniqueRowNo))
                            blnfound = True
                            break
                        if blnList!=True:
                            if counter == 1:
                                UniqueRowNo = r
                                #print("Unique Row number is : " + str(UniqueRowNo))
                                blnfound = True
                                break
        if blnfound :
            for getcolnum in range(1, maxcols + 1):
                if sheet.cell(row=1, column=getcolnum).value == strValue:
                    #print("Uniquecolvalue = " + str(getcolnum))
                    break
            c = sheet.cell(row=UniqueRowNo, column=getcolnum).value
            #print("The Value of " + strValue + " based on Unique Column and Row Values is : " + str(c))
            return c



def CheckIfDefinedElementExistInRepo(type,name):
    idtype = ReadDataFromRepo(type,name,"locator")
    idvalue = ReadDataFromRepo(type,name,"locatorvalue")
    return idtype,idvalue

def getDataForDynamicAlgo(type):
    locator = ReadDataFromJsonFile(type,"locator")
    locatorvalue = ReadDataFromJsonFile(type,"locatorvalue")
    return locator,locatorvalue